import openpyxl
import pandas as pd

class SpreadSheet:
    def __init__(self, spreadsheet_file):
        self.spreadsheet_file = spreadsheet_file
        self.wb = openpyxl.load_workbook(self.spreadsheet_file)
        self.data = list()
        self.header_name = list()

    def sheetTabSelect(self, tab_name, header_found=True):
        worksheet = self.wb[tab_name]

        if not header_found:
            for index, row in enumerate(worksheet.iter_rows()):
                if index == 0:
                    for cell in row:
                        self.header_name.append(cell.value)
                break

        for index, row in enumerate(worksheet.iter_rows()):
            row_data = list()
            if index != 0:
                for cell in row:
                    row_data.append(cell.value)
                self.data.append(row_data)



class SpreadSheet2DataFrame(SpreadSheet):
    def __init__(self, spreadsheet_file):
        super(SpreadSheet2DataFrame, self).__init__(spreadsheet_file)

    def converter(self):
        self.sheetTabSelect('Problemas', header_found=False)
        self.sheetTabSelect('Dados')
        self._df = pd.DataFrame(self.data, columns=self.header_name)

    def getDataFrame(self):
        return self._df