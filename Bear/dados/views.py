import base64
import logging
from io import BytesIO

import openpyxl
from django.contrib import messages
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse

from Clustering import KMeansMethod
import pandas as pd


import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np
from matplotlib.backends.backend_agg import FigureCanvasAgg
from django.http import HttpResponse
import io

def view_alpha(request):
    data_to_cluster = []

    columns_name = ['PROVEDOR', 'TIPO_ATIVO', 'INDEXADOR', 'QTDE.', 'RATING_AQUISICAO', 'rating_atual', 'EMISSOR',
                    'CODIGO', 'TAXA_EMISSAO', 'TAXA_AQUISICAO', 'PORC_INDEX_EMISSAO', 'PORC_INDEX_AQUISICAO',
                    'PU_AQUISICAO', 'PU_EMISSAO', 'DATA_EMISSAO', 'DATA_AQUISICAO', 'DATA_VENCIMENTO', 'NOME_MITRA',
                    'DU']

    df_to_cluster = pd.DataFrame(data_to_cluster, columns=columns_name)

    graphic = None

    return render(request, 'textos_alpha.html', {'graphic': graphic, 'tables': None})


    # template = loader.get_template('textos_alpha.html')
    # context = {
    # }
    # return HttpResponse(template.render(context, request))

def upload_spreadsheet_file(request):
    data = {}
    if "GET" == request.method:
        return HttpResponseRedirect(reverse("dados:home"))
    # if not GET, then proceed
    try:
        spreadsheet_file = request.FILES["spreadsheet_file"]
        if (not spreadsheet_file.name.endswith('.xlsx')):
            messages.error(request, 'File is not XLSX type')
            return HttpResponseRedirect(reverse("dados:home"))
        # if file is too large, return
        wb = openpyxl.load_workbook(spreadsheet_file)

        # getting a particular sheet by name out of many sheets

        data_to_cluster = list()
        columns_name = list()
        # iterating over the rows and
        # getting value from each cell in row
        worksheet = wb["Problemas"]
        for index, row in enumerate(worksheet.iter_rows()):
            row_data = list()
            if index == 0:
                for cell in row:
                    columns_name.append(cell.value)
            else:
                for cell in row:
                    row_data.append(cell.value)
                data_to_cluster.append(row_data)

        worksheet = wb["Dados"]
        for index, row in enumerate(worksheet.iter_rows()):
            row_data = list()
            if index != 0:
                for cell in row:
                    row_data.append(cell.value)
                data_to_cluster.append(row_data)

        df_to_cluster = pd.DataFrame(data_to_cluster, columns=columns_name)
        rating_existentes = df_to_cluster['rating_atual'].drop_duplicates().dropna()

        result = KMeansMethod(df_to_cluster[['DU', 'TAXA_AQUISICAO']], n_clusters=rating_existentes.shape[0])
        result.startProcessing()
        tables = list()
        df_to_cluster['Clusters'] = result.get_clusters()
        for cluster in df_to_cluster['Clusters'].unique():
            tables.append(df_to_cluster[df_to_cluster['Clusters'] == cluster].transpose() .to_dict())

        plt = result.get_plt()
        plt.tight_layout()

        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()

        graphic = base64.b64encode(image_png)
        graphic = graphic.decode('utf-8')

        return render(request, 'textos_alpha.html', {'graphic': graphic, 'tables': tables})

    except Exception as e:
        logging.getLogger("error_logger").error("Unable to upload file. " + repr(e))
        messages.error(request, "Unable to upload file. " + repr(e))

    return HttpResponseRedirect(reverse("dados:home"))